import os
import openpyxl
from PyQt5.QtCore import QSettings
from PyQt5.QtWidgets import *
from openpyxl.styles import Side, Border, Alignment
from veriçekme_python import Ui_MainWindow
import sys
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import time
import pandas as pd


class fiyatfarki(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui=Ui_MainWindow()
        self.ui.setupUi(self)
        self.getsetting() #setting çağrılır.

        url = "https://www.maliyetbul.com/tuik-fiyat-farkinda-kullanilan-endeksler-yi-ufe.php"
        html = requests.get(url).content
        self.soup = BeautifulSoup(html, "html.parser")

        # olayların tetiklenmesi
        self.ui.btn_vericek.clicked.connect(self.veri_cek)
        self.ui.btn_fiyatfark.clicked.connect(self.hesaplama)
        self.ui.open_excel.clicked.connect(self.excel_ac)
        self.ui.actionAyarlar.triggered.connect(self.showDialog)

        # Settings ayarları 3 form açıldığından kaydedilen değer çağrılır lineedite
        self.katsayi_value = self.setting_value.value('text box')
        self.ui.lineEdit.setText(self.katsayi_value)


        # BeautifulSoup ile tablodaki yılları comboboxlara çekiyoruz
        list = self.soup.find("select", {"class":"sm-form-control sx-form-control"}).findAll("option")
        for li in list:
           yil=li.text.strip().strip('Yılı')
           self.ui.cmb_ilk_yil.addItem(yil)
           self.ui.cmb_tml_yil.addItem(yil)
        try:
            self.ui.cmb_tml_yil.currentTextChanged.connect(self.aylar)
        except:
            pass
        self.aylar()


    # region Program settings ayarları
    def getsetting(self):
        # textbox
        self.setting_value=QSettings('Set_App','variables')

    def closeEvent(self, event):
        # textbox
        self.setting_value.setValue('text box',self.ui.lineEdit.text())
    # endregion

    def showDialog(self):
        text, ok = QInputDialog.getText(self, 'Fiyat Farkı Katsayısı', 'Fiyat Farkı Katsayısını giriniz(B)?')
        if ok:
            self.ui.lineEdit.setText(str(text))

    def excel_ac(self):
        os.system('Fiyat_Farkı_Hesabi.xlsx')




    def aylar(self):
        self.ui.cmb_tml_ay.clear()
        self.browser = webdriver.Chrome('chromedriver.exe')
        self.browser.get("https://www.maliyetbul.com/tuik-fiyat-farkinda-kullanilan-endeksler-yi-ufe.php")
        time.sleep(2)
        self.browser.find_element_by_xpath("//*[@id='D2']").send_keys(str(self.ui.cmb_tml_yil.currentText()))
        time.sleep(2)

        index1 = self.browser.find_element_by_xpath("//*[@id='gridbox']/div[2]/table/tbody").find_elements_by_tag_name("tr")
        endeksler1 = []
        indeksler1 = []
        for i in index1:
            endeksler1.append(i.text.split("-"))
        endeksler1.pop(0)
        for idx in endeksler1:
            indeksler1.append(idx[1].split())
        for a in indeksler1:
            self.ui.cmb_tml_ay.addItem(a[0])

        self.browser.close()

    def temel_endeks(self):
        self.browser = webdriver.Chrome('chromedriver.exe')
        self.browser.get("https://www.maliyetbul.com/tuik-fiyat-farkinda-kullanilan-endeksler-yi-ufe.php")
        time.sleep(2)
        self.browser.find_element_by_xpath("//*[@id='D2']").send_keys(str(self.ui.cmb_tml_yil.currentText()))
        time.sleep(2)
        index = self.browser.find_element_by_xpath("//*[@id='gridbox']/div[2]/table/tbody").find_elements_by_tag_name(
            "tr")
        endeksler1 = []
        indeksler1 = []
        yillar1 = []

        for i in index:
            endeksler1.append(i.text.split("-"))
        endeksler1.pop(0)
        for idx in endeksler1:
            indeksler1.append(idx[1].split())
            yillar1.append(idx[0])
        df = pd.DataFrame(indeksler1, columns=["Ay", "Güncel Endeks"])
        self.browser.close()
        print(df)
        ay=self.ui.cmb_tml_ay.currentText()
        df2=df[df["Ay"] == ay]
        self.temel_e=df2.values.tolist()


    def veri_cek(self):
        self.browser = webdriver.Chrome('chromedriver.exe')
        self.browser.get("https://www.maliyetbul.com/tuik-fiyat-farkinda-kullanilan-endeksler-yi-ufe.php")
        time.sleep(2)
        self.browser.find_element_by_xpath("//*[@id='D2']").send_keys(str(self.ui.cmb_ilk_yil.currentText()))
        time.sleep(2)
        index=self.browser.find_element_by_xpath("//*[@id='gridbox']/div[2]/table/tbody").find_elements_by_tag_name("tr")
        endeksler=[]
        indeksler=[]
        yillar=[]
        ödenekler=[]
        Katsayi=[]
        Hesap=[]
        Fiyat_Farkı=[]
        genel_endeks=[]
        temels_endeks=[]
        for i in index:
            endeksler.append(i.text.split("-"))
        endeksler.pop(0)
        for idx in endeksler:
            indeksler.append(idx[1].split())
            yillar.append(idx[0])
            ödenekler.append("0,00")
            Katsayi.append(self.ui.lineEdit.text())
            Hesap.append("0,00")
            Fiyat_Farkı.append("0,00")
            genel_endeks.append("0,00")
        df=pd.DataFrame(indeksler,columns = ["Ay","Güncel Endeks"])
        df.index.names=["İndex"]
        df["Ödenek Miktarı"]=ödenekler
        df["Fiyat Farkı Katsayısı"]=Katsayi
        df["Hesaplar"]=Hesap
        df["Fiyat Farkı"]=Fiyat_Farkı
        self.browser.close()
        self.temel_endeks()
        for idx1 in endeksler:
            temels_endeks.append(self.temel_e[0][1])
        df["Temel Endeks"]=temels_endeks
        df=pd.DataFrame.reindex(df, columns=["Ay","Güncel Endeks","Temel Endeks","Ödenek Miktarı","Fiyat Farkı Katsayısı","Hesaplar","Fiyat Farkı"])
        df.to_excel("veriler.xlsx")

        # Excel verilerini QtableWidget e çekmek
        self.df1 = pd.read_excel("veriler.xlsx")
        liste = []
        for col in self.df1.columns:
            liste.append(col)
        self.ui.tbl_fiyat_farki.setColumnCount(len(liste))
        self.ui.tbl_fiyat_farki.setHorizontalHeaderLabels(liste)
        self.ui.tbl_fiyat_farki.setColumnCount(len(self.df1.columns))
        self.ui.tbl_fiyat_farki.setRowCount(len(self.df1.index))

        for i in range(len(self.df1.index)):
            for j in range(len(self.df1.columns)):
                self.ui.tbl_fiyat_farki.setItem(i, j, QTableWidgetItem(str(self.df1.iat[i, j])))
                self.ui.tbl_fiyat_farki.horizontalHeader().setVisible(True)

        self.ui.tbl_fiyat_farki.resizeColumnsToContents()
        self.ui.tbl_fiyat_farki.resizeRowsToContents()
        self.statusBar().showMessage("Dosya başarı ile açıldı.")

    # qtablewidget de tablo üzerinde fiyat farkı hesabı yapılıyor.
    def hesaplama(self):
        rows = self.ui.tbl_fiyat_farki.rowCount()
        for i in range(rows):
            odenek = float(self.ui.tbl_fiyat_farki.item(i, 4).text().replace(",", "."))
            katsayi = float(self.ui.tbl_fiyat_farki.item(i, 5).text().replace(",", "."))
            GE = float(self.ui.tbl_fiyat_farki.item(i, 2).text().replace(",", "."))
            TE = float(self.ui.tbl_fiyat_farki.item(i, 3).text().replace(",", "."))
            FF = round(odenek * katsayi * ((GE / TE) - 1), 2)
            hsp=str(odenek)+" * "+str(katsayi)+ " * " +" (( "+str(GE)+ " / " +str(TE)+" ) - 1) "
            self.ui.tbl_fiyat_farki.setItem(i, 6, QTableWidgetItem(str(hsp)))
            self.ui.tbl_fiyat_farki.setItem(i, 7, QTableWidgetItem(str(FF)))

        self.ff_excel_aktarimi()

    def ff_excel_aktarimi(self):
        columnHeaders = []

        # create column header list
        for j in range(self.ui.tbl_fiyat_farki.columnCount()):
            columnHeaders.append(self.ui.tbl_fiyat_farki.horizontalHeaderItem(j).text())

        df3 = pd.DataFrame(columns=columnHeaders)

        # create dataframe object recordset
        for row in range(self.ui.tbl_fiyat_farki.rowCount()):
            for col in range(self.ui.tbl_fiyat_farki.columnCount()):
                item = self.ui.tbl_fiyat_farki.item(row, col)
                df3.at[row, columnHeaders[col]] = item.text() if item is not None else ""

        df3.to_excel('Fiyat_Farkı_Hesabi.xlsx', index=False)
        filename = 'Fiyat_Farkı_Hesabi.xlsx'

        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb['Sheet1']

        # region excel dosyasının biçim ayarları
        # sütun genişlikleri
        ws.column_dimensions['a'].width = 10
        # ws.column_dimensions['b'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['d'].width = 15
        ws.column_dimensions['e'].width = 15
        ws.column_dimensions['f'].width = 20
        ws.column_dimensions['g'].width = 45
        ws.column_dimensions['h'].width = 10

        # border ekleme
        thin = Side(border_style="thin", color="000000")  # Border style, color
        border = Border(left=thin, right=thin, top=thin, bottom=thin)  # Position of border
        num = self.ui.tbl_fiyat_farki.rowCount()
        for row in ws["A1:H" + str(num + 2)]:
            for cell in row:
                cell.border = border  # A5:D6 range cell setting border

        # başlık isimlerinin değiştirilmesi
        ws["A1"]="S.N"
        ws["C1"]="Güncel Endeks\n ("+str(self.ui.cmb_ilk_yil.currentText())+")"
        ws["D1"]="Temel Endeks\n ("+str(self.ui.cmb_tml_yil.currentText())+"-"+str(self.ui.cmb_tml_ay.currentText())+")"
        ws["E1"]="Ödenek Miktarı\n(A)"
        ws["F1"]="Fiyat Farkı Katsayısı\n(B)"
        ws["G1"]="Hesaplar\n(Po=Güncel Endeks/Temel Endeks)\n(F=A*B*(Po-1))"

        # hücre içindeki verilerin hizalanması
        ws["A1"].alignment = Alignment(wrap_text=True,horizontal='center')
        ws["B1"].alignment = Alignment(wrap_text=True,horizontal='center')
        ws["C1"].alignment = Alignment(wrap_text=True,horizontal='center')
        ws["D1"].alignment = Alignment(wrap_text=True,horizontal='center')
        ws["G1"].alignment = Alignment(wrap_text=True,horizontal='center')
        ws["F1"].alignment = Alignment(wrap_text=True,horizontal='center')
        ws["E1"].alignment = Alignment(wrap_text=True,horizontal='center')
        ws["H1"].alignment = Alignment(wrap_text=True,horizontal='center')
        # hücre birleştirme ve isim atama
        ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=7)
        ws["A8"] = "TOPLAM"
        ws['A8'].alignment = Alignment(horizontal='right', vertical='center')
        # endregion

        # region exceldeki belli bir sütunundaki değerleri toplama
        Ff=[]
        total=0.00
        for i in range(2,num + 2):
            a=float(ws["H"+str(i)].value.replace(",", "."))
            Ff.append(a)
        for ele in range(0, len(Ff)):
            total = total + Ff[ele]
        ws["H8"] = str(round(total,2))+" TL"
        # endregion

        # dosyayı kaydetme
        wb.save('Fiyat_Farkı_Hesabi.xlsx')
        self.statusBar().showMessage("Excele aktarım başarılı.")


if __name__=="__main__":
    app=QApplication(sys.argv)
    pencere=fiyatfarki()
    pencere.show()
    sys.exit(app.exec())